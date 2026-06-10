import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws1 = wb.active
ws1.title = "Phân tích món ăn (Detailed)"

# Styling
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# --- Sheet 1: Dish Analysis ---
headers = ["Nhóm Chi phí", "Hạng mục", "Đơn giá/Định lượng", "Thành tiền (VNĐ)", "% Doanh thu"]
ws1.append(["BẢNG PHÂN TÍCH CHI PHÍ CHI TIẾT TỪNG MÓN ĂN (Mika Casa)"])
ws1.merge_cells('A1:E1')
ws1['A1'].font = Font(size=14, bold=True)
ws1['A1'].alignment = center_align

# Example 1: Bò Steak Á-Âu
ws1.append(["MÓN 1: BÒ STEAK Á-ÂU (SIGNATURE)", "", "", "Giá bán niêm yết:", 550000])
ws1.append(headers)

data_steak = [
    ["Cost 1 (Food)", "Thịt bò thăn ngoại (250g)", "400,000/kg", 100000, "18.2%"],
    ["", "Khoai tây & Rau củ", "Set", 15000, "2.7%"],
    ["", "Sốt & Gia vị đặc biệt", "Set", 20000, "3.6%"],
    ["", "Decor & Phụ liệu", "Set", 10000, "1.8%"],
    ["Cost 2 (Vận hành)", "Lương nhân viên (Phân bổ)", "15%", 82500, "15.0%"],
    ["", "Điện, Nước, Gas", "4%", 22000, "4.0%"],
    ["", "Mặt bằng & Khấu hao", "6%", 33000, "6.0%"],
    ["", "Hoa hồng Đối tác (Vận tải)", "15%", 82500, "15.0%"],
    ["", "Marketing & Khác", "5%", 27500, "5.0%"],
    ["TỔNG CỘNG", "", "", 392500, "71.4%"],
    ["LỢI NHUẬN RÒNG", "", "", 157500, "28.6%"]
]

for row in data_steak:
    ws1.append(row)

# Spacing
ws1.append([])

# Example 2: Cocktail Mika Special
ws1.append(["MÓN 2: COCKTAIL MIKA SPECIAL", "", "", "Giá bán niêm yết:", 180000])
ws1.append(headers)
data_cocktail = [
    ["Cost 1 (Food)", "Rượu nền (Gin/Vodka)", "600,000/chai", 25000, "13.9%"],
    ["", "Trái cây tươi & Syrup", "Set", 8000, "4.4%"],
    ["", "Đá viên tinh khiết", "Set", 2000, "1.1%"],
    ["Cost 2 (Vận hành)", "Tổng định phí & Hoa hồng", "45%", 81000, "45.0%"],
    ["TỔNG CỘNG", "", "", 116000, "64.4%"],
    ["LỢI NHUẬN RÒNG", "", "", 64000, "35.6%"]
]
for row in data_cocktail:
    ws1.append(row)

# Formatting Sheet 1
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for cell in row:
        cell.border = border
        if cell.value in headers or "MÓN" in str(cell.value):
            cell.fill = sub_header_fill
            cell.font = bold_font

# --- Sheet 2: Cost 2 Calculation ---
ws2 = wb.create_sheet("Cơ cấu Cost 2 (Dự kiến)")
ws2.append(["BẢNG PHÂN BỔ CHI PHÍ VẬN HÀNH (COST 2)"])
ws2.append(["Mục tiêu doanh thu tháng:", 1000000000, "VNĐ"])
ws2.append([])
ws2.append(["Hạng mục", "Chi phí tháng (Dự kiến)", "Tỷ lệ %/Doanh thu", "Ghi chú"])

c2_data = [
    ["Tiền thuê mặt bằng", 50000000, "5.0%", "Khu vực An Thượng"],
    ["Lương nhân sự (Full team)", 150000000, "15.0%", "Bếp, Phục vụ, Sale, BA"],
    ["Điện, Nước, Internet, Gas", 40000000, "4.0%", "Vận hành 1000m2"],
    ["Hoa hồng Đối tác (Vận tải/Hotel)", 150000000, "15.0%", "Trung bình 10-20%"],
    ["Marketing & Quảng bá", 30000000, "3.0%", "Landing Page, ADS"],
    ["Khấu hao & Sửa chữa", 30000000, "3.0%", "Duy trì cơ sở vật chất"],
    ["TỔNG CHI PHÍ VẬN HÀNH (COST 2)", 450000000, "45.0%", "Mức an toàn cho F&B"]
]
for row in c2_data:
    ws2.append(row)

# --- Sheet 3: Menu Matrix ---
ws3 = wb.create_sheet("Ma trận Menu (Marketing)")
ws3.append(["MA TRẬN KỸ THUẬT THỰC ĐƠN (MENU ENGINEERING MATRIX)"])
ws3.append(["Dựa trên 3 ngày chạy thử"])
ws3.append([])
matrix_headers = ["Tên món ăn", "Số lượng bán", "Giá bán", "Tổng Cost (1+2)", "Lợi nhuận/Món", "Phân loại Ma trận", "Hành động (Action)"]
ws3.append(matrix_headers)

matrix_data = [
    ["Bò Steak Á-Âu", 45, 550000, 392500, 157500, "STARS", "Đẩy mạnh quảng bá"],
    ["Mì Ý Hải Sản", 60, 250000, 180000, 70000, "PLOW HORSES", "Xem lại giá hoặc cost"],
    ["Cocktail Mika", 80, 180000, 116000, 64000, "STARS", "Món 'mồi' tốt"],
    ["Gỏi cuốn Tôm Thịt", 15, 120000, 95000, 25000, "DOGS", "Gỡ khỏi Menu ngay"],
    ["Rượu Vang Đỏ", 10, 1200000, 700000, 500000, "PUZZLES", "Cần Sale chào khách"]
]
for row in matrix_data:
    ws3.append(row)

# Styling for all sheets
for sheet in [ws1, ws2, ws3]:
    for col in range(1, 8):
        sheet.column_dimensions[get_column_letter(col)].width = 25

wb.save("restaurant_cost_analysis_mika_casa.xlsx")